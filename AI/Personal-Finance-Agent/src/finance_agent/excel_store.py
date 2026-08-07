from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .categorizer import KEYWORD_CATEGORIES
from .models import Transaction


TRANSACTION_HEADERS = [
    "transaction_id",
    "bank",
    "occurred_at",
    "month",
    "transaction_type",
    "amount",
    "category",
    "review_status",
    "include_in_spending",
    "counterparty",
    "content",
]

AUDIT_HEADERS = [
    "transaction_id",
    "email_id",
    "source",
    "currency",
    "sender",
    "receiver",
    "account_hint",
    "balance_after",
    "category_source",
    "raw_subject",
    "raw_from",
    "raw_snippet",
]

SUMMARY_HEADERS = ["month", "income", "expense", "net_cashflow", "transaction_count", "needs_review_count"]
CATEGORY_HEADERS = ["keyword", "category"]


class ExcelStore:
    def __init__(self, path: str | Path):
        self.path = Path(path)
        self.path.parent.mkdir(parents=True, exist_ok=True)
        created = not self.path.exists()
        self.workbook = self._load_or_create()
        self._format_transactions()
        self._format_audit()
        if created:
            self.save()

    def append_transactions(self, transactions: list[Transaction]) -> list[Transaction]:
        sheet = self.workbook["Transactions"]
        audit_sheet = self.workbook["Audit"]
        existing_ids = self.transaction_ids()
        inserted: list[Transaction] = []

        for transaction in transactions:
            if transaction.transaction_id in existing_ids:
                continue
            sheet.append(_transaction_to_row(transaction))
            audit_sheet.append(_transaction_to_audit_row(transaction))
            existing_ids.add(transaction.transaction_id)
            inserted.append(transaction)

        if inserted:
            self._format_transactions()
            self._format_audit()
            self.refresh_monthly_summary()
            self.save()
        return inserted

    def transaction_ids(self) -> set[str]:
        sheet = self.workbook["Transactions"]
        ids: set[str] = set()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                ids.add(str(row[0]))
        return ids

    def load_transactions(self) -> list[Transaction]:
        sheet = self.workbook["Transactions"]
        headers = _headers(sheet)
        transactions: list[Transaction] = []
        for values in sheet.iter_rows(min_row=2, values_only=True):
            if not values or not values[0]:
                continue
            row = dict(zip(headers, values))
            row.update(self._audit_row(str(row.get("transaction_id") or "")))
            transactions.append(_row_to_transaction(row))
        return transactions

    def update_transaction_category(
        self,
        transaction_id: str,
        category: str,
        include_in_spending: bool,
        category_source: str,
    ) -> bool:
        updated = False
        sheet = self.workbook["Transactions"]
        headers = _headers(sheet)
        columns = {header: index + 1 for index, header in enumerate(headers)}
        for row_idx in range(2, sheet.max_row + 1):
            if str(sheet.cell(row_idx, columns["transaction_id"]).value or "") != transaction_id:
                continue
            sheet.cell(row_idx, columns["category"]).value = category
            sheet.cell(row_idx, columns["review_status"]).value = "confirmed"
            sheet.cell(row_idx, columns["include_in_spending"]).value = include_in_spending
            updated = True
            break

        if "Audit" in self.workbook.sheetnames:
            audit_sheet = self.workbook["Audit"]
            audit_headers = _headers(audit_sheet)
            audit_columns = {header: index + 1 for index, header in enumerate(audit_headers)}
            if "category_source" in audit_columns:
                for row_idx in range(2, audit_sheet.max_row + 1):
                    if str(audit_sheet.cell(row_idx, audit_columns["transaction_id"]).value or "") == transaction_id:
                        audit_sheet.cell(row_idx, audit_columns["category_source"]).value = category_source
                        break

        if updated:
            self._format_transactions()
            self._format_audit()
            self.refresh_monthly_summary()
            self.save()
        return updated

    def refresh_monthly_summary(self) -> None:
        if "Monthly Summary" in self.workbook.sheetnames:
            del self.workbook["Monthly Summary"]
        sheet = self.workbook.create_sheet("Monthly Summary")
        sheet.append(SUMMARY_HEADERS)

        summary: dict[str, dict[str, float]] = {}
        counts: dict[str, int] = {}
        review_counts: dict[str, int] = {}
        for transaction in self.load_transactions():
            month = _month_key(transaction.occurred_at)
            bucket = summary.setdefault(month, {"income": 0.0, "expense": 0.0})
            if transaction.needs_review:
                review_counts[month] = review_counts.get(month, 0) + 1
            if transaction.transaction_type == "income":
                bucket["income"] += transaction.amount
            elif transaction.include_in_spending:
                bucket["expense"] += transaction.amount
            counts[month] = counts.get(month, 0) + 1

        for month in sorted(summary):
            income = summary[month]["income"]
            expense = summary[month]["expense"]
            sheet.append([month, income, expense, income - expense, counts[month], review_counts.get(month, 0)])

        self._format_sheet(sheet)

    def save(self) -> None:
        self.workbook.save(self.path)

    def _load_or_create(self) -> Workbook:
        if self.path.exists():
            workbook = load_workbook(self.path)
            self._ensure_sheets(workbook)
            return workbook

        workbook = Workbook()
        default = workbook.active
        default.title = "Transactions"
        default.append(TRANSACTION_HEADERS)
        workbook.create_sheet("Audit").append(AUDIT_HEADERS)
        workbook.create_sheet("Monthly Summary").append(SUMMARY_HEADERS)
        self._ensure_sheets(workbook)
        return workbook

    def _ensure_sheets(self, workbook: Workbook) -> None:
        if "Transactions" not in workbook.sheetnames:
            workbook.create_sheet("Transactions", 0).append(TRANSACTION_HEADERS)
        elif _headers(workbook["Transactions"]) != TRANSACTION_HEADERS:
            _migrate_transactions_to_current_schema(workbook)

        if "Audit" not in workbook.sheetnames:
            workbook.create_sheet("Audit").append(AUDIT_HEADERS)

        if "Monthly Summary" not in workbook.sheetnames:
            workbook.create_sheet("Monthly Summary").append(SUMMARY_HEADERS)

        if "Categories" not in workbook.sheetnames:
            sheet = workbook.create_sheet("Categories")
            sheet.append(CATEGORY_HEADERS)
            for keyword, category in sorted(KEYWORD_CATEGORIES.items(), key=lambda item: (item[1], item[0])):
                sheet.append([keyword, category])
            self._format_sheet(sheet)

    def _audit_row(self, transaction_id: str) -> dict[str, object]:
        if "Audit" not in self.workbook.sheetnames:
            return {}
        sheet = self.workbook["Audit"]
        headers = _headers(sheet)
        for values in sheet.iter_rows(min_row=2, values_only=True):
            row = dict(zip(headers, values))
            if str(row.get("transaction_id") or "") == transaction_id:
                return row
        return {}

    def _format_transactions(self) -> None:
        sheet = self.workbook["Transactions"]
        self._format_sheet(sheet)
        widths = {
            "A": 28,
            "B": 16,
            "C": 22,
            "D": 12,
            "E": 16,
            "F": 14,
            "G": 18,
            "H": 16,
            "I": 18,
            "J": 42,
            "K": 34,
        }
        for col, width in widths.items():
            sheet.column_dimensions[col].width = width
        for row in sheet.iter_rows(min_row=2, min_col=6, max_col=6):
            row[0].number_format = '#,##0 "VND"'
        sheet.column_dimensions["A"].hidden = True

    def _format_audit(self) -> None:
        if "Audit" not in self.workbook.sheetnames:
            return
        sheet = self.workbook["Audit"]
        self._format_sheet(sheet)
        widths = {
            "A": 28,
            "B": 28,
            "C": 12,
            "D": 12,
            "E": 28,
            "F": 42,
            "G": 16,
            "H": 16,
            "I": 28,
            "J": 32,
            "K": 28,
            "L": 64,
        }
        for col, width in widths.items():
            sheet.column_dimensions[col].width = width
        sheet.sheet_state = "hidden"

    def _format_sheet(self, sheet) -> None:
        header_fill = PatternFill("solid", fgColor="1F2937")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        sheet.freeze_panes = "A2"
        for col_idx in range(1, sheet.max_column + 1):
            letter = get_column_letter(col_idx)
            if not sheet.column_dimensions[letter].width:
                sheet.column_dimensions[letter].width = 16


def _transaction_to_row(transaction: Transaction) -> list[object]:
    return [
        transaction.transaction_id,
        transaction.bank,
        transaction.occurred_at.isoformat(sep=" ", timespec="seconds") if transaction.occurred_at else "",
        _month_key(transaction.occurred_at),
        transaction.transaction_type,
        transaction.amount,
        transaction.category,
        "needs_review" if transaction.needs_review else "confirmed",
        transaction.include_in_spending,
        transaction.receiver,
        transaction.content,
    ]


def _transaction_to_audit_row(transaction: Transaction) -> list[object]:
    return [
        transaction.transaction_id,
        transaction.email_id,
        transaction.source,
        transaction.currency,
        transaction.sender,
        transaction.receiver,
        transaction.account_hint,
        transaction.balance_after,
        transaction.category_source,
        transaction.raw_subject,
        transaction.raw_from,
        transaction.raw_snippet,
    ]


def _row_to_transaction(row: dict[str, object]) -> Transaction:
    review_status = str(row.get("review_status") or "")
    category = str(row.get("category") or "Cần phân loại")
    return Transaction(
        transaction_id=str(row.get("transaction_id") or ""),
        email_id=str(row.get("email_id") or ""),
        source=str(row.get("source") or "email"),
        bank=str(row.get("bank") or "Unknown"),
        occurred_at=_parse_datetime(row.get("occurred_at")),
        transaction_type=str(row.get("transaction_type") or "expense"),
        amount=float(row.get("amount") or 0),
        currency=str(row.get("currency") or "VND"),
        content=str(row.get("content") or ""),
        sender=str(row.get("sender") or ""),
        receiver=str(row.get("receiver") or row.get("counterparty") or ""),
        account_hint=str(row.get("account_hint") or ""),
        balance_after=float(row["balance_after"]) if row.get("balance_after") not in {None, ""} else None,
        category=category,
        raw_subject=str(row.get("raw_subject") or ""),
        raw_from=str(row.get("raw_from") or ""),
        raw_snippet=str(row.get("raw_snippet") or ""),
        category_source=str(row.get("category_source") or ""),
        needs_review=review_status == "needs_review" or category == "Cần phân loại",
        include_in_spending=_parse_bool(row.get("include_in_spending")),
    )


def _parse_datetime(value: object) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value))
    except ValueError:
        return None


def _month_key(value: datetime | None) -> str:
    value = value or datetime.now()
    return value.strftime("%Y-%m")


def _headers(sheet) -> list[object]:
    return [cell.value for cell in sheet[1]]


def _parse_bool(value: object) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    return str(value).strip().lower() in {"1", "true", "yes", "y"}


def _migrate_transactions_to_current_schema(workbook: Workbook) -> None:
    old_sheet = workbook["Transactions"]
    old_headers = _headers(old_sheet)
    rows = [
        dict(zip(old_headers, values))
        for values in old_sheet.iter_rows(min_row=2, values_only=True)
        if values and values[0]
    ]

    del workbook["Transactions"]
    sheet = workbook.create_sheet("Transactions", 0)
    sheet.append(TRANSACTION_HEADERS)

    if "Audit" in workbook.sheetnames:
        del workbook["Audit"]
    audit_sheet = workbook.create_sheet("Audit")
    audit_sheet.append(AUDIT_HEADERS)

    for row in rows:
        transaction = _row_to_transaction(row)
        sheet.append(_transaction_to_row(transaction))
        audit_sheet.append(_transaction_to_audit_row(transaction))
